#!/usr/bin/env python3
"""
Fill Товари_для_заповнення in cursor_characteristics_fill_pack.xlsx using:
- Правила_категорій (skip not_applicable attributes)
- Джерело_товарів (explicit values by SKU)
- Title parsing (Патерни_розбору) when source cell is empty

CRM import (later): add Product.characteristics JSON in Prisma, extend PATCH /products/:id,
then for each filled row map columns F–AE (except source_fragment/fill_status/review_note for
customer API) to a JSON object keyed by attribute_code and upsert by sku.
"""

from __future__ import annotations

import re
import sys
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.workbook.workbook import Workbook

# Column indices (1-based) on Товари_для_заповнення — must match sheet header row
COL = {
    "sku": 1,
    "name": 2,
    "price": 3,
    "category_name": 4,
    "subcategory_name": 5,
    "compatibility_raw": 6,
    "compatibility": 7,
    "implant_system": 8,
    "connection_type": 9,
    "platform": 10,
    "diameter": 11,
    "height": 12,
    "gingival_height": 13,
    "angle": 14,
    "length": 15,
    "material": 16,
    "coating": 17,
    "color": 18,
    "screw_included": 19,
    "screwdriver_type": 20,
    "packaging_qty": 21,
    "sterile": 22,
    "production_time": 23,
    "tray_type": 24,
    "restoration_type": 25,
    "for_multi_unit": 26,
    "position_shape": 27,
    "length_variant": 28,
    "profile_size": 29,
    "source_fragment": 30,
    "fill_status": 31,
    "review_note": 32,
}

ATTR_CODES = list(COL.keys())[5:]  # from compatibility_raw through review_note

SOURCE_HEADERS = [
    "Артикул",
    "Название",
    "Цена",
    "Категория",
    "Підкатегорія",
    "Шаблон характеристик",
    "Сумісність / система",
    "Для мульти-юніта",
    "Тип / виконання",
    "Тип реставрації",
    "Ложка",
    "Матеріал",
    "GH, мм",
    "AH, мм",
    "Діаметр, мм",
    "Кут, °",
    "Розмір / профіль",
    "Довжина / версія",
    "Позиція / форма",
    "Кількість в упаковці, шт",
    "Примітка",
]


def load_category_rules(ws) -> dict[str, set[str]]:
    """attribute_code -> allowed for category (excludes not_applicable)."""
    allowed: dict[str, set[str]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0] or not row[1]:
            continue
        cat, code, _, status = row[0], row[1], row[2], row[3]
        if status == "not_applicable":
            continue
        allowed.setdefault(str(cat).strip(), set()).add(str(code).strip())
    return allowed


def load_source_by_sku(ws) -> dict[str, dict[str, Any]]:
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    idx = {h: i for i, h in enumerate(headers) if h}
    out: dict[str, dict[str, Any]] = {}
    for r in range(2, ws.max_row + 1):
        sku = ws.cell(r, 1).value
        if not sku:
            continue
        key = str(sku).strip()
        row = {}
        for h in SOURCE_HEADERS:
            if h in idx:
                row[h] = ws.cell(r, idx[h] + 1).value
        out[key] = row
    return out


def num_val(v: Any) -> float | None:
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", ".")
    try:
        return float(s) if "." in s else float(int(s))
    except ValueError:
        return None


def bool_from_source(v: Any) -> bool | None:
    if v is None or v == "":
        return None
    s = str(v).strip().lower()
    if s in ("так", "yes", "true", "1"):
        return True
    if s in ("ні", "no", "false", "0"):
        return False
    return None


def normalize_tray(v: Any) -> str | None:
    if v is None or v == "":
        return None
    s = str(v).strip().lower()
    if "відкрит" in s:
        return "відкритої ложки"
    if "закрит" in s:
        return "закритої ложки"
    return str(v).strip()


def raw_compat_from_title(name: str) -> str | None:
    """Extract raw compatibility token when clearly present."""
    # Longer tokens first
    patterns = [
        r"MG\s*AR/AO",
        r"ST\s*RC/NC",
        r"ST\s*RC",
        r"ST\s*MU",
        r"MG\s*AR",
        r"\bNC\b",
        r"\bRC\b",
        r"\bAO\b",
    ]
    found: list[str] = []
    for pat in patterns:
        m = re.search(pat, name, re.IGNORECASE)
        if m:
            t = m.group(0).replace(" ", " ").strip()
            if t.upper() not in [x.upper() for x in found]:
                found.append(m.group(0).strip())
    if not found:
        return None
    # Prefer single combined token as in source (e.g. ST RC)
    if any("ST" in f and "RC" in name for f in found):
        if re.search(r"ST\s*RC/NC", name, re.I):
            return "ST RC/NC"
        if re.search(r"ST\s*RC", name, re.I):
            return "ST RC"
        if re.search(r"ST\s*MU", name, re.I):
            return "ST MU"
    return found[0]


def parse_title(name: str) -> dict[str, Any]:
    out: dict[str, Any] = {}
    if not name:
        return out

    m = re.search(r"GH\s*([0-9]+(?:[.,][0-9]+)?)\s*mm", name, re.IGNORECASE)
    if m:
        out["gingival_height"] = float(m.group(1).replace(",", "."))

    m = re.search(r"AH\s*([0-9]+(?:[.,][0-9]+)?)\s*mm", name, re.IGNORECASE)
    if m:
        out["height"] = float(m.group(1).replace(",", "."))

    m = re.search(r"([0-9]+(?:[.,][0-9]+)?)\s*°", name)
    if m:
        out["angle"] = float(m.group(1).replace(",", "."))

    m = re.search(r"⌀\s*([0-9]+(?:[.,][0-9]+)?)", name)
    if m:
        out["diameter"] = float(m.group(1).replace(",", "."))

    m = re.search(r"(\d+)\s*шт", name, re.IGNORECASE)
    if m:
        out["packaging_qty"] = int(m.group(1))

    if re.search(r"коронка", name, re.IGNORECASE):
        out["restoration_type"] = "коронка"
    elif re.search(r"\bміст\b", name, re.IGNORECASE) or re.search(r"мост", name, re.IGNORECASE):
        out["restoration_type"] = "міст"

    if re.search(r"відкритої\s+ложки|відкрита\s+ложка", name, re.IGNORECASE):
        out["tray_type"] = "відкритої ложки"
    elif re.search(r"закритої\s+ложки|закрита\s+ложка", name, re.IGNORECASE):
        out["tray_type"] = "закритої ложки"

    mtrx = re.search(r"TRX\s+(L21|S16)", name, re.IGNORECASE)
    if mtrx:
        out["length_variant"] = mtrx.group(1).upper()
    else:
        for lv in ("Long", "Short", "S16", "L21"):
            if re.search(rf"\b{re.escape(lv)}\b", name):
                out["length_variant"] = lv
                break

    # Explicit multi-unit product (not "ST MU" compatibility code)
    if re.search(r"мульти[- ]?юніт|мультиюніт|кутового\s+мульти", name, re.IGNORECASE):
        out["for_multi_unit"] = True

    m = re.search(r"(?:GH[0-9]+(?:[.,][0-9]+)?mm)\s+([A-Z])\s*$", name, re.IGNORECASE)
    if m:
        out["position_shape"] = m.group(1).upper()
    elif re.search(r"Мульти|кутовий", name, re.IGNORECASE):
        m = re.search(r"\s([A-C])\s*$", name)
        if m:
            out["position_shape"] = m.group(1).upper()

    # profile sizes: M1.4 on multi-unit holder, bit sizes 1.20, platform 3.5/3.75
    if re.search(r"MU\s+M\s*([0-9]+(?:[.,][0-9]+)?)", name, re.IGNORECASE):
        m = re.search(r"\bM\s*([0-9]+(?:[.,][0-9]+)?)\b", name, re.IGNORECASE)
        if m:
            out["profile_size"] = f"M{m.group(1)}".replace(",", ".")
    elif "Біта" in name:
        m = re.search(r"SUPREX\s+([0-9]+[.,][0-9]+)", name, re.IGNORECASE)
        if m:
            out["profile_size"] = m.group(1).replace(",", ".")
    if "Титанова" in name and "платформа" in name:
        m2 = re.search(r"(\d+\s*/\s*\d+(?:[.,]\d+)?)", name)
        if m2:
            out["profile_size"] = m2.group(1).replace(" ", "")

    if re.search(r"Титанова\s+платформа", name) or re.search(r"\bТитан\b", name):
        out["material"] = "Титан"
    m = re.search(r"Co[- ]?Cr|Сo[- ]?Cr", name, re.IGNORECASE)
    if m:
        out["material"] = "Co-Cr"

    if re.search(r"з\s+гвинтом|гвинт\s+у\s+комплекті", name, re.IGNORECASE):
        out["screw_included"] = True

    if "SUPREX" in name:
        out["screwdriver_type"] = "SUPREX"

    rc = raw_compat_from_title(name)
    if rc:
        out["compatibility_raw"] = rc

    return out


def merge_source_into_attrs(src: dict[str, Any]) -> dict[str, Any]:
    m: dict[str, Any] = {}
    v = src.get("Сумісність / система")
    if v not in (None, ""):
        m["compatibility_raw"] = str(v).strip()

    b = bool_from_source(src.get("Для мульти-юніта"))
    if b is True:
        m["for_multi_unit"] = True

    v = src.get("Тип реставрації")
    if v not in (None, ""):
        m["restoration_type"] = str(v).strip()

    tt = normalize_tray(src.get("Ложка"))
    if tt:
        m["tray_type"] = tt

    v = src.get("Матеріал")
    if v not in (None, ""):
        m["material"] = str(v).strip()

    g = num_val(src.get("GH, мм"))
    if g is not None:
        m["gingival_height"] = g

    h = num_val(src.get("AH, мм"))
    if h is not None:
        m["height"] = h

    d = num_val(src.get("Діаметр, мм"))
    if d is not None:
        m["diameter"] = d

    a = num_val(src.get("Кут, °"))
    if a is not None:
        m["angle"] = a

    v = src.get("Розмір / профіль")
    if v not in (None, ""):
        m["profile_size"] = str(v).strip()

    v = src.get("Довжина / версія")
    if v not in (None, ""):
        m["length_variant"] = str(v).strip()

    v = src.get("Позиція / форма")
    if v not in (None, ""):
        m["position_shape"] = str(v).strip()

    pq = num_val(src.get("Кількість в упаковці, шт"))
    if pq is not None:
        m["packaging_qty"] = int(pq)

    return m


def coalesce_bool(v: Any) -> bool | None:
    if v is True or v is False:
        return v
    return None


def fill_workbook(path: Path) -> None:
    wb: Workbook = openpyxl.load_workbook(path, data_only=False)
    rules_ws = wb["Правила_категорій"]
    allowed_by_cat = load_category_rules(rules_ws)
    source = load_source_by_sku(wb["Джерело_товарів"])
    tgt = wb["Товари_для_заповнення"]

    for row_idx in range(2, tgt.max_row + 1):
        sku_c = tgt.cell(row_idx, COL["sku"]).value
        if not sku_c:
            cat_only = tgt.cell(row_idx, COL["category_name"]).value
            if cat_only and str(cat_only).strip() == "Новий товар":
                tgt.cell(row_idx, COL["fill_status"]).value = "review_needed"
                tgt.cell(row_idx, COL["review_note"]).value = "немає SKU в рядку"
            elif tgt.cell(row_idx, COL["name"]).value:
                tgt.cell(row_idx, COL["fill_status"]).value = "review_needed"
                tgt.cell(row_idx, COL["review_note"]).value = "немає SKU в рядку"
            continue
        sku = str(sku_c).strip()
        cat_cell = tgt.cell(row_idx, COL["category_name"]).value
        category = str(cat_cell).strip() if cat_cell else ""
        allowed = allowed_by_cat.get(category, set())

        src_row = source.get(sku, {})
        name_cell = tgt.cell(row_idx, COL["name"]).value
        name = str(name_cell) if name_cell else ""

        merged: dict[str, Any] = {}
        merged.update(parse_title(name))
        # Source overrides / supplements explicit spreadsheet fields
        merged.update(merge_source_into_attrs(src_row))

        # Preserve existing non-empty values where we do not have a new value
        for key in ATTR_CODES:
            if key in ("source_fragment", "fill_status", "review_note"):
                continue
            col = COL[key]
            existing = tgt.cell(row_idx, col).value
            if existing not in (None, "") and key not in merged:
                if key in (
                    "screw_included",
                    "sterile",
                    "for_multi_unit",
                ) and isinstance(existing, str):
                    if existing.upper() == "TRUE":
                        merged[key] = True
                    elif existing.upper() == "FALSE":
                        merged[key] = False
                    else:
                        merged[key] = existing
                else:
                    merged[key] = existing

        fragments: list[str] = []

        # Apply category filter and write
        filled = 0
        review_bits: list[str] = []

        for key in ATTR_CODES:
            if key in ("source_fragment", "fill_status", "review_note"):
                continue
            if key not in allowed:
                tgt.cell(row_idx, COL[key]).value = None
                continue

            val = merged.get(key)
            if val in (None, ""):
                continue

            # Type coercion
            if key in ("screw_included", "sterile", "for_multi_unit"):
                b = coalesce_bool(val)
                if b is None:
                    continue
                tgt.cell(row_idx, COL[key]).value = bool(b)
                filled += 1
                continue

            if key in (
                "diameter",
                "height",
                "gingival_height",
                "angle",
                "length",
                "packaging_qty",
                "production_time",
            ):
                n = num_val(val)
                if n is None:
                    continue
                if key == "packaging_qty":
                    tgt.cell(row_idx, COL[key]).value = int(n)
                else:
                    tgt.cell(row_idx, COL[key]).value = n
                filled += 1
                continue

            tgt.cell(row_idx, COL[key]).value = val
            filled += 1

        # source_fragment
        if src_row:
            fragments.append("Джерело_товарів")
        if name:
            fragments.append("назва")
        frag_text = " + ".join(fragments) if fragments else ""
        if frag_text:
            tgt.cell(row_idx, COL["source_fragment"]).value = frag_text

        # fill_status
        if category == "Новий товар" or (category and not allowed):
            status = "review_needed"
            review_bits.append("категорія без правил або Новий товар")
        elif filled == 0:
            status = "untouched"
        else:
            status = "partial" if filled < 4 else "done"

        tgt.cell(row_idx, COL["fill_status"]).value = status
        if review_bits:
            tgt.cell(row_idx, COL["review_note"]).value = "; ".join(review_bits)

    wb.save(path)


def main() -> None:
    p = Path(sys.argv[1]) if len(sys.argv) > 1 else Path("/Users/konstantin/Downloads/cursor_characteristics_fill_pack.xlsx")
    if not p.is_file():
        print("File not found:", p, file=sys.stderr)
        sys.exit(1)
    fill_workbook(p)
    print("Saved:", p)


if __name__ == "__main__":
    main()
